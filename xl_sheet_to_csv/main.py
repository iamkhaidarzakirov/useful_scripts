from openpyxl import load_workbook
import pandas as pd
# Getting all sheets from Excel file. Converting to csv format and saving each sheet as new file
wb = load_workbook('data.xlsx')
sheets_titles = wb.sheetnames
print(sheets_titles)
for title in sheets_titles:
    df = pd.read_excel('data.xlsx', sheet_name=f'{title}')
    df.to_csv(f'{title}', header=True)
